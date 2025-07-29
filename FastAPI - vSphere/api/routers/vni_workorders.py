from fastapi import APIRouter, HTTPException, Depends, Path
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from app.database import get_db
from models.vni_workorder import VNIWorkOrder
from services.vsphere.vni_operations import VNIOperations
from datetime import datetime
import json
import os
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter(
    prefix="/vni-workorders",
    tags=["VNI WorkOrders"]
)

@router.post("/")
def create_vni_workorder(
    vni_workorder: dict,
    db: Session = Depends(get_db)
):
    """Create a new VNI workorder"""
    print("Received VNI workorder:", vni_workorder)
    try:
        # Parse dates
        requested_date = vni_workorder.get("requested_date")
        if requested_date:
            try:
                requested_date = datetime.fromisoformat(requested_date)
            except Exception:
                requested_date = datetime.utcnow()
        else:
            requested_date = datetime.utcnow()
            
        deadline = vni_workorder.get("deadline")
        if deadline:
            try:
                deadline = datetime.fromisoformat(deadline)
            except Exception:
                deadline = datetime.utcnow()
        else:
            deadline = datetime.utcnow()
        
        new_vni_order = VNIWorkOrder(
            owner=vni_workorder["owner"],
            requested_date=requested_date,
            requested_by=vni_workorder["requested_by"],
            virtual_machines=vni_workorder.get("virtual_machines", []),
            deadline=deadline,
            project=vni_workorder["project"],
            t0_gw=vni_workorder["t0_gw"],
            t1_gw=vni_workorder["t1_gw"],
            description=vni_workorder["description"],
            vni_name=vni_workorder["vni_name"],
            cidr=vni_workorder["cidr"],
            subnet_mask=vni_workorder["subnet_mask"],
            gateway=vni_workorder["gateway"],
            first_ip=vni_workorder["first_ip"],
            last_ip=vni_workorder["last_ip"],
            number_of_ips=vni_workorder["number_of_ips"],
            status="pending",
            notes=vni_workorder.get("notes"),
            priority=vni_workorder.get("priority", "normal"),
            assigned_to=vni_workorder.get("assigned_to")
        )
        
        db.add(new_vni_order)
        db.commit()
        db.refresh(new_vni_order)
        
        return {
            "id": new_vni_order.id,
            "owner": new_vni_order.owner,
            "requested_date": new_vni_order.requested_date.isoformat(),
            "requested_by": new_vni_order.requested_by,
            "virtual_machines": new_vni_order.virtual_machines,
            "deadline": new_vni_order.deadline.isoformat(),
            "project": new_vni_order.project,
            "t0_gw": new_vni_order.t0_gw,
            "t1_gw": new_vni_order.t1_gw,
            "description": new_vni_order.description,
            "vni_name": new_vni_order.vni_name,
            "cidr": new_vni_order.cidr,
            "subnet_mask": new_vni_order.subnet_mask,
            "gateway": new_vni_order.gateway,
            "first_ip": new_vni_order.first_ip,
            "last_ip": new_vni_order.last_ip,
            "number_of_ips": new_vni_order.number_of_ips,
            "status": new_vni_order.status,
            "created_at": new_vni_order.created_at.isoformat(),
            "notes": new_vni_order.notes,
            "priority": new_vni_order.priority,
            "assigned_to": new_vni_order.assigned_to
        }
    except Exception as e:
        import traceback
        print('CREATE VNI WORKORDER ERROR:', e)
        traceback.print_exc()
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/")
def get_vni_workorders(
    skip: int = 0,
    limit: int = 100,
    status: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Get all VNI workorders with optional filtering"""
    try:
        query = db.query(VNIWorkOrder)
        
        if status:
            query = query.filter(VNIWorkOrder.status == status)
            
        vni_workorders = query.offset(skip).limit(limit).all()
        
        return [
            {
                "id": wo.id,
                "owner": wo.owner,
                "requested_date": wo.requested_date.isoformat(),
                "requested_by": wo.requested_by,
                "virtual_machines": wo.virtual_machines,
                "deadline": wo.deadline.isoformat(),
                "project": wo.project,
                "t0_gw": wo.t0_gw,
                "t1_gw": wo.t1_gw,
                "description": wo.description,
                "vni_name": wo.vni_name,
                "cidr": wo.cidr,
                "subnet_mask": wo.subnet_mask,
                "gateway": wo.gateway,
                "first_ip": wo.first_ip,
                "last_ip": wo.last_ip,
                "number_of_ips": wo.number_of_ips,
                "status": wo.status,
                "created_at": wo.created_at.isoformat(),
                "updated_at": wo.updated_at.isoformat(),
                "notes": wo.notes,
                "priority": wo.priority,
                "assigned_to": wo.assigned_to
            }
            for wo in vni_workorders
        ]
    except Exception as e:
        print('GET VNI WORKORDERS ERROR:', e)
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/{vni_workorder_id}")
def get_vni_workorder(
    vni_workorder_id: int = Path(..., description="The ID of the VNI workorder"),
    db: Session = Depends(get_db)
):
    """Get a specific VNI workorder by ID"""
    try:
        vni_workorder = db.query(VNIWorkOrder).filter(VNIWorkOrder.id == vni_workorder_id).first()
        if not vni_workorder:
            raise HTTPException(status_code=404, detail="VNI workorder not found")
            
        return {
            "id": vni_workorder.id,
            "owner": vni_workorder.owner,
            "requested_date": vni_workorder.requested_date.isoformat(),
            "requested_by": vni_workorder.requested_by,
            "virtual_machines": vni_workorder.virtual_machines,
            "deadline": vni_workorder.deadline.isoformat(),
            "project": vni_workorder.project,
            "t0_gw": vni_workorder.t0_gw,
            "t1_gw": vni_workorder.t1_gw,
            "description": vni_workorder.description,
            "vni_name": vni_workorder.vni_name,
            "cidr": vni_workorder.cidr,
            "subnet_mask": vni_workorder.subnet_mask,
            "gateway": vni_workorder.gateway,
            "first_ip": vni_workorder.first_ip,
            "last_ip": vni_workorder.last_ip,
            "number_of_ips": vni_workorder.number_of_ips,
            "status": vni_workorder.status,
            "created_at": vni_workorder.created_at.isoformat(),
            "updated_at": vni_workorder.updated_at.isoformat(),
            "last_execution_log": vni_workorder.last_execution_log,
            "notes": vni_workorder.notes,
            "priority": vni_workorder.priority,
            "assigned_to": vni_workorder.assigned_to
        }
    except HTTPException:
        raise
    except Exception as e:
        print('GET VNI WORKORDER ERROR:', e)
        raise HTTPException(status_code=500, detail=str(e))

@router.put("/{vni_workorder_id}")
def update_vni_workorder(
    vni_workorder_id: int,
    vni_workorder_update: dict,
    db: Session = Depends(get_db)
):
    """Update a VNI workorder"""
    try:
        vni_workorder = db.query(VNIWorkOrder).filter(VNIWorkOrder.id == vni_workorder_id).first()
        if not vni_workorder:
            raise HTTPException(status_code=404, detail="VNI workorder not found")
        
        # Update fields
        for field, value in vni_workorder_update.items():
            if hasattr(vni_workorder, field):
                if field in ["requested_date", "deadline"] and value:
                    try:
                        value = datetime.fromisoformat(value)
                    except Exception:
                        value = datetime.utcnow()
                setattr(vni_workorder, field, value)
        
        vni_workorder.updated_at = datetime.utcnow()
        db.commit()
        db.refresh(vni_workorder)
        
        return {
            "id": vni_workorder.id,
            "owner": vni_workorder.owner,
            "requested_date": vni_workorder.requested_date.isoformat(),
            "requested_by": vni_workorder.requested_by,
            "virtual_machines": vni_workorder.virtual_machines,
            "deadline": vni_workorder.deadline.isoformat(),
            "project": vni_workorder.project,
            "t0_gw": vni_workorder.t0_gw,
            "t1_gw": vni_workorder.t1_gw,
            "description": vni_workorder.description,
            "vni_name": vni_workorder.vni_name,
            "cidr": vni_workorder.cidr,
            "subnet_mask": vni_workorder.subnet_mask,
            "gateway": vni_workorder.gateway,
            "first_ip": vni_workorder.first_ip,
            "last_ip": vni_workorder.last_ip,
            "number_of_ips": vni_workorder.number_of_ips,
            "status": vni_workorder.status,
            "created_at": vni_workorder.created_at.isoformat(),
            "updated_at": vni_workorder.updated_at.isoformat(),
            "notes": vni_workorder.notes,
            "priority": vni_workorder.priority,
            "assigned_to": vni_workorder.assigned_to
        }
    except HTTPException:
        raise
    except Exception as e:
        print('UPDATE VNI WORKORDER ERROR:', e)
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/{vni_workorder_id}/approve")
def approve_vni_workorder(
    vni_workorder_id: int,
    db: Session = Depends(get_db)
):
    """Approve a VNI workorder"""
    try:
        vni_workorder = db.query(VNIWorkOrder).filter(VNIWorkOrder.id == vni_workorder_id).first()
        if not vni_workorder:
            raise HTTPException(status_code=404, detail="VNI workorder not found")
        
        vni_workorder.status = "approved"
        vni_workorder.updated_at = datetime.utcnow()
        db.commit()
        
        return {"message": "VNI workorder approved successfully", "status": "approved"}
    except HTTPException:
        raise
    except Exception as e:
        print('APPROVE VNI WORKORDER ERROR:', e)
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/{vni_workorder_id}/reject")
def reject_vni_workorder(
    vni_workorder_id: int,
    db: Session = Depends(get_db)
):
    """Reject a VNI workorder"""
    try:
        vni_workorder = db.query(VNIWorkOrder).filter(VNIWorkOrder.id == vni_workorder_id).first()
        if not vni_workorder:
            raise HTTPException(status_code=404, detail="VNI workorder not found")
        
        vni_workorder.status = "rejected"
        vni_workorder.updated_at = datetime.utcnow()
        db.commit()
        
        return {"message": "VNI workorder rejected successfully", "status": "rejected"}
    except HTTPException:
        raise
    except Exception as e:
        print('REJECT VNI WORKORDER ERROR:', e)
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@router.put("/{vni_workorder_id}/status")
def update_vni_workorder_status(
    vni_workorder_id: int,
    status_update: dict,
    db: Session = Depends(get_db)
):
    """Update VNI workorder status"""
    try:
        vni_workorder = db.query(VNIWorkOrder).filter(VNIWorkOrder.id == vni_workorder_id).first()
        if not vni_workorder:
            raise HTTPException(status_code=404, detail="VNI workorder not found")
        
        new_status = status_update.get("status")
        valid_statuses = ["pending", "approved", "rejected", "draft", "executing", "completed", "failed"]
        
        if new_status not in valid_statuses:
            raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {', '.join(valid_statuses)}")
        
        vni_workorder.status = new_status
        vni_workorder.updated_at = datetime.utcnow()
        db.commit()
        
        return {"message": f"VNI workorder status updated to {new_status}", "status": new_status}
    except HTTPException:
        raise
    except Exception as e:
        print('UPDATE VNI WORKORDER STATUS ERROR:', e)
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/{vni_workorder_id}/execute")
def execute_vni_workorder(
    vni_workorder_id: int,
    db: Session = Depends(get_db)
):
    """Execute a VNI workorder using the VNI operations service"""
    try:
        vni_workorder = db.query(VNIWorkOrder).filter(VNIWorkOrder.id == vni_workorder_id).first()
        if not vni_workorder:
            raise HTTPException(status_code=404, detail="VNI workorder not found")
        
        if vni_workorder.status != "approved":
            raise HTTPException(status_code=400, detail="VNI workorder must be approved before execution")
        
        # Update status to executing
        vni_workorder.status = "executing"
        vni_workorder.updated_at = datetime.utcnow()
        vni_workorder.last_execution_log = f"Execution started at {datetime.utcnow().isoformat()}"
        db.commit()
        
        # Use VNI operations service
        vni_ops = VNIOperations()
        
        # Prepare VNI configuration
        vni_config = {
            "vni_name": vni_workorder.vni_name,
            "cidr": vni_workorder.cidr,
            "gateway": vni_workorder.gateway,
            "t0_gw": vni_workorder.t0_gw,
            "t1_gw": vni_workorder.t1_gw,
            "description": vni_workorder.description,
            "project": vni_workorder.project,
            "subnet_mask": vni_workorder.subnet_mask,
            "first_ip": vni_workorder.first_ip,
            "last_ip": vni_workorder.last_ip,
            "number_of_ips": vni_workorder.number_of_ips
        }
        
        # Validate configuration
        validation_result = vni_ops.validate_vni_config(vni_config)
        if not validation_result["valid"]:
            vni_workorder.status = "failed"
            vni_workorder.last_execution_log = f"Validation failed: {', '.join(validation_result['errors'])}"
            vni_workorder.updated_at = datetime.utcnow()
            db.commit()
            raise HTTPException(status_code=400, detail=f"VNI configuration validation failed: {', '.join(validation_result['errors'])}")
        
        # Create VNI
        result = vni_ops.create_vni(vni_config)
        
        if result["success"]:
            vni_workorder.status = "completed"
            vni_workorder.last_execution_log = f"VNI '{vni_workorder.vni_name}' created successfully at {datetime.utcnow().isoformat()}. VNI ID: {result.get('vni_id', 'N/A')}"
        else:
            vni_workorder.status = "failed"
            vni_workorder.last_execution_log = f"VNI creation failed: {result.get('message', 'Unknown error')}"
        
        vni_workorder.updated_at = datetime.utcnow()
        db.commit()
        
        return {
            "message": result.get("message", "VNI workorder execution completed"),
            "status": vni_workorder.status,
            "vni_name": vni_workorder.vni_name,
            "vni_id": result.get("vni_id"),
            "execution_log": vni_workorder.last_execution_log
        }
    except HTTPException:
        raise
    except Exception as e:
        print('EXECUTE VNI WORKORDER ERROR:', e)
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@router.delete("/{vni_workorder_id}")
def delete_vni_workorder(
    vni_workorder_id: int,
    db: Session = Depends(get_db)
):
    """Delete a VNI workorder"""
    try:
        vni_workorder = db.query(VNIWorkOrder).filter(VNIWorkOrder.id == vni_workorder_id).first()
        if not vni_workorder:
            raise HTTPException(status_code=404, detail="VNI workorder not found")
        
        db.delete(vni_workorder)
        db.commit()
        
        return {"message": "VNI workorder deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        print('DELETE VNI WORKORDER ERROR:', e)
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/{vni_workorder_id}/log")
def get_vni_workorder_log(
    vni_workorder_id: int,
    db: Session = Depends(get_db)
):
    """Get the execution log for a VNI workorder"""
    try:
        vni_workorder = db.query(VNIWorkOrder).filter(VNIWorkOrder.id == vni_workorder_id).first()
        if not vni_workorder:
            raise HTTPException(status_code=404, detail="VNI workorder not found")
        
        return {"log": vni_workorder.last_execution_log}
    except HTTPException:
        raise
    except Exception as e:
        print('GET VNI WORKORDER LOG ERROR:', e)
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/{vni_workorder_id}/status")
def get_vni_workorder_status(
    vni_workorder_id: int,
    db: Session = Depends(get_db)
):
    """Get the status of a VNI workorder"""
    try:
        vni_workorder = db.query(VNIWorkOrder).filter(VNIWorkOrder.id == vni_workorder_id).first()
        if not vni_workorder:
            raise HTTPException(status_code=404, detail="VNI workorder not found")
        
        return {"status": vni_workorder.status}
    except HTTPException:
        raise
    except Exception as e:
        print('GET VNI WORKORDER STATUS ERROR:', e)
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/{vni_workorder_id}/export-excel")
def export_vni_workorder_excel(
    vni_workorder_id: int,
    db: Session = Depends(get_db)
):
    """Export VNI workorder to Excel with Ooredoo styling"""
    try:
        vni_workorder = db.query(VNIWorkOrder).filter(VNIWorkOrder.id == vni_workorder_id).first()
        if not vni_workorder:
            raise HTTPException(status_code=404, detail="VNI workorder not found")
        
        # Create a new workbook and select the active sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "VNI Configuration"
        
        # Define colors based on the image (red and white theme)
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        red_font = Font(color="FF0000", bold=True, size=14)
        white_font = Font(color="FFFFFF", bold=True, size=12)
        black_font = Font(color="000000", size=11)
        bold_black_font = Font(color="000000", bold=True, size=11)
        
        # Define borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Center alignment
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        
        # Row 1: Ooredoo logo and header info
        ws.merge_cells('A1:K1')
        ooredoo_cell = ws['A1']
        ooredoo_cell.value = "ooredoo"
        ooredoo_cell.font = red_font
        ooredoo_cell.alignment = left_alignment
        
        # Row 2-6: Information block (Owner, Requested Date, etc.)
        info_labels = ["Owner:", "Requested Date:", "Requested By:", "Virtual Machines:", "Dead Line :"]
        info_values = [
            vni_workorder.owner,
            vni_workorder.requested_date.strftime("%m/%d/%Y") if vni_workorder.requested_date else "",
            vni_workorder.requested_by,
            str(len(vni_workorder.virtual_machines)) if vni_workorder.virtual_machines else "0",
            vni_workorder.deadline.strftime("%m/%d/%Y") if vni_workorder.deadline else ""
        ]
        
        for i, (label, value) in enumerate(zip(info_labels, info_values)):
            row = i + 2
            # Label cell (red background, white text)
            label_cell = ws[f'A{row}']
            label_cell.value = label
            label_cell.fill = red_fill
            label_cell.font = white_font
            label_cell.alignment = left_alignment
            label_cell.border = thin_border
            
            # Value cell (white background, black text)
            value_cell = ws[f'B{row}']
            value_cell.value = value
            value_cell.fill = white_fill
            value_cell.font = black_font
            value_cell.alignment = left_alignment
            value_cell.border = thin_border
        
        # Row 7: Main title "Création VNI"
        ws.merge_cells('A7:K7')
        title_cell = ws['A7']
        title_cell.value = "Création VNI"
        title_cell.fill = red_fill
        title_cell.font = white_font
        title_cell.alignment = center_alignment
        
        # Row 8: Table headers
        headers = ["Projet", "t0-gw", "t1-gw", "Description", "vni name", "CIDR", "Masque", "gateway", "first Ip", "last IP", "number"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col)
            cell.value = header
            cell.fill = red_fill
            cell.font = white_font
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # Row 9: Data row
        data_values = [
            vni_workorder.project,
            vni_workorder.t0_gw,
            vni_workorder.t1_gw,
            vni_workorder.description,
            vni_workorder.vni_name,
            vni_workorder.cidr,
            vni_workorder.subnet_mask,
            vni_workorder.gateway,
            vni_workorder.first_ip,
            vni_workorder.last_ip,
            str(vni_workorder.number_of_ips)
        ]
        
        for col, value in enumerate(data_values, 1):
            cell = ws.cell(row=9, column=col)
            cell.value = value
            cell.fill = white_fill
            cell.font = black_font
            cell.alignment = center_alignment
            cell.border = thin_border
            
            # Make gateway bold
            if col == 8:  # gateway column
                cell.font = bold_black_font
        
        # Set column widths
        column_widths = [15, 12, 12, 15, 40, 15, 15, 15, 12, 12, 8]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Create temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            wb.save(tmp_file.name)
            tmp_file_path = tmp_file.name
        
        # Return the file
        filename = f"VNI_Configuration_{vni_workorder.vni_name}_{vni_workorder.id}.xlsx"
        return FileResponse(
            path=tmp_file_path,
            filename=filename,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'}
        )
    except HTTPException:
        raise
    except Exception as e:
        print('EXPORT VNI WORKORDER EXCEL ERROR:', e)
        raise HTTPException(status_code=500, detail=str(e)) 